"""Install openpyxl then read the HOA ledger XLS."""
import subprocess, sys, os

# Install
subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "xlrd", "-q"], capture_output=True)

XLS = r"C:\Users\andre\Desktop\COURT_FILING_PACKETS\SHADY\shady enhanced redacted ledger$$$ conv_10aa496d.xls"

if not os.path.exists(XLS):
    print(f"FILE NOT FOUND: {XLS}")
    sys.exit(1)

# Try xlrd first (old .xls format)
try:
    import xlrd
    wb = xlrd.open_workbook(XLS)
    print(f"xlrd opened: {wb.nsheets} sheets")
    for si in range(wb.nsheets):
        sh = wb.sheet_by_index(si)
        print(f"\n=== Sheet {si}: {sh.name} ({sh.nrows} rows x {sh.ncols} cols) ===")
        for ri in range(min(sh.nrows, 60)):
            row = [str(sh.cell_value(ri, ci)) for ci in range(sh.ncols)]
            line = " | ".join(c for c in row if c.strip())
            if line.strip():
                print(f"  Row {ri}: {line}")
except Exception as e:
    print(f"xlrd failed: {e}")
    # Try openpyxl
    try:
        import openpyxl
        wb = openpyxl.load_workbook(XLS, data_only=True)
        print(f"openpyxl opened: {wb.sheetnames}")
        for sn in wb.sheetnames:
            ws = wb[sn]
            print(f"\n=== Sheet: {sn} ===")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 60: break
                vals = [str(v) for v in row if v is not None]
                if vals:
                    print(f"  {' | '.join(vals)}")
    except Exception as e2:
        print(f"openpyxl also failed: {e2}")
