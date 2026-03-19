from __future__ import annotations
import argparse, csv, json, hashlib
from pathlib import Path
from datetime import datetime, timedelta

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
IMPORTS = ROOT / "imports" / "docket"

def load_json(path: Path, default):
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))

def save_json(path: Path, payload):
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")

def parse_date(s):
    s = str(s or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d","%m/%d/%Y","%m/%d/%y","%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            continue
    # pass-through if already close
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10]
    return None

def read_rows(path: Path):
    suf = path.suffix.lower()
    if suf in {".csv",".tsv"}:
        delim = "\t" if suf == ".tsv" else ","
        with path.open("r", encoding="utf-8", errors="ignore", newline="") as f:
            return list(csv.DictReader(f, delimiter=delim))
    if suf == ".json":
        obj = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(obj, list):
            return obj
        if isinstance(obj, dict):
            for k in ("rows","entries","items","docket"):
                if isinstance(obj.get(k), list):
                    return obj[k]
        return []
    if suf in {".xlsx",".xlsm"}:
        try:
            from openpyxl import load_workbook  # type: ignore
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(h or "").strip() for h in rows[0]]
            out = []
            for r in rows[1:]:
                out.append({headers[i] if i < len(headers) else f"col{i}": r[i] for i in range(len(r))})
            return out
        except Exception:
            return []
    return []

def normalize_entry(row: dict, source_path: str, idx: int):
    norm = {str(k).strip().lower(): v for k,v in row.items()}
    def pick(*keys):
        for k in keys:
            if k in norm and norm[k] not in (None,""):
                return norm[k]
        return None
    date_val = parse_date(pick("date","entry_date","filed_date","entered_date","hearing_date","event_date"))
    title = str(pick("title","text","description","entry","docket_text","event") or "").strip()
    case_id = str(pick("case_id","case number","case_number","case") or "").strip()
    kind = str(pick("kind","type","entry_type","event_type") or "").strip() or ("hearing" if "hearing" in title.lower() else "order" if "order" in title.lower() else "docket_entry")
    entry_id = str(pick("entry_id","id","doc_id","document_id") or "").strip()
    if not entry_id:
        seed = f"{case_id}|{date_val}|{title}|{kind}|{source_path}|{idx}"
        entry_id = "IM-" + hashlib.md5(seed.encode()).hexdigest()[:10]
    return {
        "entry_id": entry_id,
        "kind": kind,
        "case_id": case_id or None,
        "date": date_val,
        "title": title or f"Imported entry {idx+1}",
        "lane": None,
        "source_ref": source_path,
        "raw_row": row,
        "provenance_jump_hints": [f"{source_path}#{idx+2}", entry_id],
    }

def infer_deadline_candidates(entries: list[dict]):
    out = []
    for e in entries:
        dt = e.get("date")
        if not dt:
            continue
        try:
            d = datetime.strptime(dt, "%Y-%m-%d").date()
        except Exception:
            continue
        title_low = str(e.get("title","")).lower()
        if "hearing" in title_low:
            for days, kind in [(21,"prep_packet"),(7,"witness_check"),(1,"appearance_check")]:
                due = (d - timedelta(days=days)).isoformat()
                out.append({
                    "deadline_id": f"IMP-{e['entry_id']}-{kind}",
                    "case_id": e.get("case_id"),
                    "source_entry_id": e["entry_id"],
                    "due_date_iso": due,
                    "title": f"{kind.replace('_',' ').title()} for {e['title']}",
                    "status": "watch",
                    "deadline_kind": kind,
                    "truth_tag": "INFERRED",
                    "basis": "Imported hearing entry scheduling checkpoint (workflow-generated)",
                    "provenance_jump_hints": e.get("provenance_jump_hints", []),
                })
        if "ex parte" in title_low or ("order" in title_low and "objection" not in title_low):
            due = (d + timedelta(days=14)).isoformat()
            out.append({
                "deadline_id": f"IMP-{e['entry_id']}-order-review",
                "case_id": e.get("case_id"),
                "source_entry_id": e["entry_id"],
                "due_date_iso": due,
                "title": f"Order review / objection window check (manual authority confirm) • {e['title']}",
                "status": "manual_authority_confirm",
                "deadline_kind": "order_review_window",
                "truth_tag": "INFERRED",
                "basis": "Workflow reminder generated from imported order-like docket text; verify exact Michigan timing rule before filing.",
                "provenance_jump_hints": e.get("provenance_jump_hints", []),
            })
    return out

def main() -> int:
    parser = argparse.ArgumentParser(description="Import MiFILE/ROA docket export files and propagate deltas + deadlines")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    files = []
    if IMPORTS.exists():
        for p in sorted(IMPORTS.rglob("*")):
            if p.is_file() and p.suffix.lower() in {".csv",".tsv",".json",".xlsx",".xlsm"}:
                files.append(p)

    imported_entries = []
    file_runs = []
    for fp in files:
        rows = read_rows(fp)
        normed = [normalize_entry(r, str(fp.relative_to(ROOT)).replace("\\","/"), i) for i, r in enumerate(rows)]
        imported_entries.extend(normed)
        file_runs.append({
            "file_path": str(fp.relative_to(ROOT)).replace("\\","/"),
            "row_count": len(rows),
            "normalized_count": len(normed),
            "status": "ok",
        })

    # Deduplicate imported entries by entry_id; keep last occurrence
    emap = {e["entry_id"]: e for e in imported_entries}
    imported_entries = sorted(emap.values(), key=lambda e: (str(e.get("date") or ""), str(e.get("entry_id"))))

    current = load_json(DATA / "docket_snapshot_current.json", {"entries":[]})
    c_entries = current.get("entries", [])
    c_map = {e.get("entry_id"): e for e in c_entries if e.get("entry_id")}
    i_map = {e.get("entry_id"): e for e in imported_entries}

    added = [i_map[k] for k in sorted(i_map.keys() - c_map.keys())]
    removed = [c_map[k] for k in sorted(c_map.keys() - i_map.keys())]
    changed = []
    for k in sorted(i_map.keys() & c_map.keys()):
        a, b = c_map[k], i_map[k]
        fp_a = json.dumps({x:a.get(x) for x in ('date','title','kind','case_id')}, sort_keys=True)
        fp_b = json.dumps({x:b.get(x) for x in ('date','title','kind','case_id')}, sort_keys=True)
        if fp_a != fp_b:
            changed.append({"entry_id": k, "before": a, "after": b})

    deadline_candidates = infer_deadline_candidates(imported_entries)

    import_runs = load_json(DATA / "live_docket_import_runs.json", {"runs":[]})
    run_id = "LDR-" + datetime.now().strftime("%Y%m%d-%H%M%S")
    run_row = {
        "run_id": run_id,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "files_processed": file_runs,
        "imported_entry_count": len(imported_entries),
        "delta_counts": {"added": len(added), "removed": len(removed), "changed": len(changed)},
        "deadline_candidate_count": len(deadline_candidates),
        "resolution_targets": [
            "Export fresh MiFILE/ROA CSV and place in imports/docket for real delta sync.",
            "Review imported deadlines marked INFERRED before routing them into filing packets."
        ],
    }
    import_runs.setdefault("runs", []).append(run_row)
    import_runs["runs"] = import_runs["runs"][-30:]
    import_runs["latest"] = run_row

    imported_snapshot = {
        "generated_at": run_row["generated_at"],
        "snapshot_id": run_id,
        "entries": imported_entries,
        "source_files": [fr["file_path"] for fr in file_runs],
    }
    delta_payload = {
        "generated_at": run_row["generated_at"],
        "sync_status": "IMPORT_READY_NO_FILES" if not files else ("BASELINE_SYNCED" if not (added or removed or changed) else "DELTAS_PENDING_REVIEW"),
        "delta_counts": run_row["delta_counts"],
        "import_snapshot_id": run_id,
        "baseline_snapshot_id": current.get("snapshot_id"),
        "added": added,
        "removed": removed,
        "changed": changed,
        "resolution_target": "Use imported ROA/MiFILE exports to drive live procedural delta review and downstream deadline rails.",
    }
    deadline_payload = {
        "generated_at": run_row["generated_at"],
        "summary": {
            "imported_entries": len(imported_entries),
            "deadline_candidates": len(deadline_candidates),
            "hearing_checkpoints": sum(1 for d in deadline_candidates if d.get("deadline_kind") in ("prep_packet","witness_check","appearance_check")),
            "order_review_windows": sum(1 for d in deadline_candidates if d.get("deadline_kind") == "order_review_window"),
        },
        "items": deadline_candidates[:400],
        "resolution_targets": [
            "Promote only deadlines with verified Michigan authority timing into the main deadlines rail.",
            "Keep workflow-generated reminders visible even when legal deadline is not yet confirmed."
        ]
    }

    if not args.dry_run:
        save_json(DATA / "live_docket_import_runs.json", import_runs)
        save_json(DATA / "docket_snapshot_imported_live.json", imported_snapshot)
        save_json(DATA / "docket_deadline_propagation.json", deadline_payload)
        # Save a sidecar delta for UI and optionally merge a live-import block into docket_delta_sync
        save_json(DATA / "docket_import_delta.json", delta_payload)
        base_sync = load_json(DATA / "docket_delta_sync.json", {})
        if isinstance(base_sync, dict):
            base_sync["live_import"] = {
                "sync_status": delta_payload["sync_status"],
                "delta_counts": delta_payload["delta_counts"],
                "import_snapshot_id": delta_payload["import_snapshot_id"],
                "resolution_target": delta_payload["resolution_target"],
            }
            save_json(DATA / "docket_delta_sync.json", base_sync)
    print(json.dumps({"status":"ok","files":len(files),"entries":len(imported_entries),"deadline_candidates":len(deadline_candidates)}, indent=2))
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
